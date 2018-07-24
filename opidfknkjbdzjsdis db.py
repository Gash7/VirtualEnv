import openpyxl
import statistics as stats

book = openpyxl.load_workbook('D:\Python\Ash.xlsx', data_only=True)

sheet = book.active

rows = sheet.rows

values = []

for row in rows:
    for cell in row:
        values.append(cell.value)

print("Number of values: {0}".format(len(values)))
#print("Sum of values: {0}".format(sum(values)))
#print("Minimum value: {0}".format(min(values)))
#print("Maximum value: {0}".format(max(values)))
print("Mean: {0}".format(stats.mean(values)))
print("Median: {0}".format(stats.median(values)))
print("Standard deviation: {0}".format(stats.stdev(values)))
print("Variance: {0}".format(stats.variance(values)))

'''
from openpyxl import Workbook

book = Workbook()
sheet = book.active

rows = (
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

for row in rows:
    sheet.append(row)

for row in sheet.iter_cols(min_row=1, min_col=1, max_row=6, max_col=3):
    for cell in row:
        print(cell.value, end=" ")
    print()

book.save('iterbycols.xlsx')



from openpyxl import Workbook

book = Workbook()
sheet = book.active

rows = (
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

for row in rows:
    sheet.append(row)
 for row in sheet.add_image('')

 for row in sheet.iter_rows(min_row=1, min_col=1, max_row=6, max_col=3):
    for cell in row:
        print(cell.value, end=" ")
    print()

book.save('iterbyrows.xlsx')





import openpyxl

book = openpyxl.load_workbook('D:\Python\Ash.xlsx')

sheet = book.active

cells = sheet['A1': 'B8']

for c1, c2 in cells:
    print("{0:9} {1:9}".format(c1.value, c2.value))



import openpyxl

book = openpyxl.load_workbook('D:\Python\Ash1.xlsx')

sheet = book.active

a1 = sheet['A1']
a2 = sheet['A2']
a3 = sheet.cell(row=3, column=1)
a4 = sheet.cell(row=2,column=20,value=1000)

print(a1.value)
print(a2.value)
print(a3.value)
print(a4.value)




from openpyxl import Workbook

book = Workbook()
sheet = book.active

rows = (
    ('Sr.', 'User Name', 'Password','Expected Message'),
    ('01', 'Ashu022', 12345,'Ok All right'),
    (2, 'Ashu023', 12344,'Ok All right'),
    (3, 'Ashu024', 12343,'Ok All right'),
    (4, 'Ashu025', 12344,'Ok All right')

)

for row in rows:
    sheet.append(row)

book.save('appending.xlsx')



from openpyxl import Workbook

book = Workbook()
sheet = book.active

sheet['A1'] = 1
sheet.cell(row=2, column=2).value = 2

book.save('write2cell.xlsx')



#from openpyxl import Workbook
import time

book = Workbook()
sheet = book.active

sheet['A1'] = 56
sheet['A2'] = 43

now = time.strftime("%x")
sheet['A3'] = now

book.save('D:\Python\Ash1.xlsx')
'''